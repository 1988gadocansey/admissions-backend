import datetime
import decimal
import os
import re
from datetime import date
from datetime import datetime
from ftplib import FTP
import pysftp
import dateutil
from django.conf import settings
import openpyxl
import paramiko
import requests
# from rest_framework.response import Response
import simplejson
# from ajax_search.forms import SearchForm
import xlsxwriter

# from decouple import config
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files.storage import FileSystemStorage
from django.core.mail import EmailMessage
from django.core.paginator import EmptyPage, PageNotAnInteger
from django.core.paginator import InvalidPage
from django.core.paginator import Paginator
from django.db.models import Count
from django.db.models import Q
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import get_template
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.chart import (
    Reference,
    BarChart
)
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter
from xlsxwriter.workbook import Workbook

from .models import Applicantmodel, Smsmodel, Resultuploadmodel, Programmemodel, Hallmodel, Regionmodel, \
    Departmentmodel, \
    Facultymodel, Countrymodel, Aspnetusers, Formerschoolmodel, Sale, Configurations
from .utils import render_to_pdf


def try_excel(request):
    return HttpResponse('hi')


# from rest_framework import status

# api for ttu.edu.gh page for applicant searching

def excel(request):
    setting = Configurations.objects.order_by('-id')[0]
    year = setting.working_year
    print(year)
    data = Applicantmodel.objects.filter(admitted=True).filter(yearofadmission=year).all()
    # Create an new Excel file and add a worksheet.
    print(data.count)
    workbook = xlsxwriter.Workbook('excel/demo.xlsx')

    workbook.add_format({'bold': True})

    applicants = workbook.add_worksheet('applicants')

    programs = workbook.add_worksheet('programs')

    # Widen the first column to make the text clearer.
    applicants.set_column('A:A', 20)
    applicants.set_column('C:C', 20)

    columnHeaders = ['Name', 'Gender', 'Programme Admitted']

    # pass a dictionary of formating to headers
    header_border = workbook.add_format(
        {
            "bottom": 6,  # top, left, right
            "bottom_color": "#ff000",
            "top": 1,  # bottom
            "top_color": "#ff0000",
            "font_size": 14

        }

    )

    rowIndex = 1
    startColumn = 0
    endColumn = startColumn + len(columnHeaders)

    # # creating headers/columns for excel
    #
    # for columIndex in range(startColumn, endColumn):
    #     applicants.write(
    #
    #         rowIndex,
    #         columIndex,
    #         columnHeaders[columIndex - startColumn],
    #
    #         header_border
    #     )

    # writing data to excel i.e rows

    for records in data:
        applicants.write(

            rowIndex,
            startColumn,
            records.lastname + records.firstname
        )

        applicants.write(

            rowIndex,
            startColumn + 1,
            records.gender
        )
        applicants.write(

            rowIndex,
            startColumn + 2,
            str(records.programmeadmittedid.name),

        )
        rowIndex += 1

    # Add a bold format to use to highlight cells.

    # write  the title o
    applicants.write("A1", "Name")
    applicants.write("B1", "Gender")
    applicants.write("C1", "Programme")

    workbook.close()
    return HttpResponse('excel created')


def set_border(ws, cell_range):
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    rows = ws.iter_rows(cell_range)
    for row in rows:
        for cell in row:
            cell.border = border


@login_required
def show_export_form(request):
    try:

        programme = Programmemodel.objects.filter(runing=1).filter(~Q(type="BTECH")).filter(
            ~Q(type="POSTGRADUATE")).order_by('name').all()

        year = ["2021/2022", "2020/2021", "2019/2020", "2018/2019", "2017/2018", "2016/2017", "2015/2016"]

        context = {'programme': programme, 'years': year}

        return render(request, 'reports/verification.html', context)
    except:

        messages.warning(request,
                         "Error displaying Applicant ")


def remove_duplicates(values):
    output = []
    seen = set()
    for value in values:
        # If value has not been encountered yet,
        # ... add it to both list and set.
        if value not in seen:
            output.append(value)
            seen.add(value)
    return output


def totalAdmittedByYear(year, program):
    data = Applicantmodel.objects.filter(yearofadmission=year).filter(admitted=True).filter(
        programmeadmittedid=program).count()

    return data


@login_required
def download_excel_report(request):
    year = request.POST.get('year')
    programme = request.POST.get('programme')

    programmeData = Programmemodel.objects.all().order_by("type")

    if str(programmeData) != "":

        # print(programmeData.query)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = "attachment; filename=Admissions_statistics.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active

        set_border(ws, "A4:D181")

        style_1 = Font(color='0000FF00', bold=True, size=15)

        ws.title = "Admissions_statistics"

        ws['A1'] = "TAKORADI TECHNICAL UNIVERSITY"
        ws['A2'] = "ADMISSIONS"
        ws['A3'] = "ADMISSIONS STATISTICS"

        row_num = 3

        columns = [
            (u"PROGRAMME", 60,),
            (u"2017/2018", 30),
            (u"2018/2019", 30),
            (u"2019/2020", 30),
            (u"2020/2021", 30),
            (u"2021/2022", 30),

        ]

        monthExam = ""
        yearExam = ""

        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            # c.style.font.bold = True
            # set column width
            ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

        for obj in programmeData:
            row_num += 1

            dataSet = [
                obj.lastname + obj.firstname + obj.middlename,
                totalAdmittedByYear("2017/2018", obj.id),
                totalAdmittedByYear("2018/2019", obj.id),
                totalAdmittedByYear("2019/2020", obj.id),
                totalAdmittedByYear("2020/2021", obj.id),
                totalAdmittedByYear("2021/2022", obj.id),

            ]
            # remove_duplicates(dataSet)
            for col_num in range(len(dataSet)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)

                c.value = dataSet[col_num]
                # c.style.alignment.wrap_text = True
                # c.style.font=style_1

        # formating

        sheet = wb.get_sheet_by_name("Admissions_statistics")
        sheet.sheet_properties.tabColor = "0072BA"

        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")

        border = Border(top=double, left=thin, right=thin, bottom=double)
        fill = PatternFill("solid", fgColor="DDDDDD")
        fill = GradientFill(stop=("000000", "FFFFFF"))
        font = Font(b=True, color="FF0000")
        al = Alignment(horizontal="center", vertical="center")

        wb.save(response)

        export_xlsx.short_description = u"Export results for verification XLSX"
        return response

    else:

        return render(request, 'old/dashboard/index.html')


@login_required
def export_xlsx(request):
    year = request.POST.get('year')
    programme = request.POST.get('programme')

    programmeData = Programmemodel.objects.filter(id=programme).get()

    programmeName = programmeData.name

    if year != "" and programme != "":

        # data = AdmissionsExamsResults.objects.filter(applicantData__year_of_admission=year).filter(applicantData__admited=1).filter(applicantData__programme_admitted__id=programme).order_by('applicantData__name').values('index_no','applicantData__name','month','applicant').annotate(total=Count('month'))

        data = Resultuploadmodel.objects.filter(applicant__yearofadmission=year).filter(applicant__admited=True).filter(
            applicant__programmeadmittedid=programme).order_by('applicant__lastname').values('index_no',
                                                                                             'applicant__lastname',
                                                                                             'year',
                                                                                             'month',
                                                                                             'form').annotate(
            total=Count('month'))
        print(data.query)
        # distinct = AdmissionsExamsResults.objects.filter(applicantData__admited=1).filter(applicantData__programme_admitted__type='HND').annotate(
        #     name_count=Count('index_no')
        # ).filter(name_count=1)
        #
        # data = AdmissionsExamsResults.objects.filter(index_no__in=[item.index_no for item in distinct])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = "attachment; filename=" + programmeName + "_ " + year + "_admitted_applicants" + ".xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active

        # ws.cell('A1:A3').border = Border(top=Side(border_style='thin', color='FF000000'),
        #                               right=Side(border_style='thin', color='FF000000'),
        #                               bottom=Side(border_style='thin', color='FF000000'),
        #                               left=Side(border_style='thin', color='FF000000'))

        style_1 = Font(color='0000FF00', bold=True, size=15)

        ws.title = "Admitted Applicants"

        ws['A1'] = "TAKORADI TECHNICAL UNIVERSITY"
        ws['A2'] = "ADMISSIONS"
        ws['A3'] = "RESULT VERIFICATION FOR " + programmeName + "- " + str(year) + " ADMISSIONS "

        row_num = 3

        columns = [
            (u"NAME", 60,),
            (u"ADMISSION NUMBER", 30),
            (u"WEAC INDEX NUMBER", 30),
            (u"YEAR OF EXAMS", 30),
            (u"MONTH", 15),
        ]

        monthExam = ""
        yearExam = ""

        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            # c.style.font.bold = True
            # set column width
            ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

        for obj in data:
            row_num += 1

            if year in ["2018/2019", "2017/2018"]:

                if "MAY/JUNE" in obj['month'] or "may/june" in obj['month'] or "june" in obj['month'] or "may" in obj[
                    'month'] or "MayJune" in obj['month'] or "March" in obj['month'] or "/June" in obj[
                    'month'] or "/May" in \
                        obj['month'] or "April" in obj['month'] or "June/" in obj['month'] or "May/" in obj[
                    'month'] or "MAY/" in obj['month'] or "JUNE/" in obj['month'] or "JUNE" in obj['month'] or "MAY" in \
                        obj[
                            'month']:
                    monthExam = 6
                else:
                    monthExam = 12

                yearExam = str(obj['month']).rpartition(' ')[2]

                if not yearExam:
                    yearExam = str(obj['month']).rpartition(',')[2]

            else:
                monthExam = obj['month']

                yearExam = obj['year']
                # if not yearExam:
            #
            #     yearExam = obj.month.rpartition(',')[2]

            dataSet = [
                obj['applicant__name'],
                obj['form'],
                obj['index_no'],
                yearExam,
                monthExam
            ]
            # remove_duplicates(dataSet)
            for col_num in range(len(dataSet)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)

                c.value = dataSet[col_num]
                # c.style.alignment.wrap_text = True
                # c.style.font=style_1

        # formating

        sheet = wb.get_sheet_by_name("Admitted Applicants")
        sheet.sheet_properties.tabColor = "0072BA"

        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")

        border = Border(top=double, left=thin, right=thin, bottom=double)
        fill = PatternFill("solid", fgColor="DDDDDD")
        fill = GradientFill(stop=("000000", "FFFFFF"))
        font = Font(b=True, color="FF0000")
        al = Alignment(horizontal="center", vertical="center")

        wb.save(response)

        export_xlsx.short_description = u"Export results for verification XLSX"
        return response

    else:

        return render(request, 'old/dashboard/index.html')


def getWeacMonth(applicant):
    data = Resultuploadmodel.objects.filter(applicant=applicant).values_list("month").annotate(
        total=Count("month"))

    return data


def graphs(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=graphs.xlsx'

    book = Workbook()
    sheet = book.active

    rows = [
        ("USA", 46),
        ("China", 38),
        ("UK", 29),
        ("Russia", 22),
        ("South Korea", 13),
        ("Germany", 11)
    ]

    for row in rows:
        sheet.append(row)

    ws2 = book.create_sheet("Graphs")

    data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=6)
    categs = Reference(sheet, min_col=1, min_row=1, max_row=6)

    chart = BarChart()
    chart.add_data(data=data)
    chart.set_categories(categs)

    chart.legend = None
    chart.y_axis.majorGridlines = None
    chart.varyColors = True
    chart.title = "Olympic Gold medals in London"

    ws2.add_chart(chart, "A8")

    # book.save("bar_chart.xlsx")

    book.save(response)
    return response


def applicantSearch(request, token):
    q = token
    applicant = Applicantmodel.objects.filter(
        Q(pin__icontains=q) | Q(programme_admitted_id__name__icontains=q) | Q(application_number__icontains=q) | Q(
            first_name__icontains=q) | Q(
            last_name__icontains=q) | Q(
            other_names__icontains=q)).filter(admited=1).filter(status="admitted").select_related(
        'programme_admitted').all()

    json_res = []
    if applicant:
        for a in applicant:
            json_obj = dict(
                name=a.name,
                programme=a.programme_admitted.name,
                hall=a.hall_admitted.hall_name,
                fees=a.admission_fees,
                firstname=a.first_name,
                lastname=a.last_name,
                age=a.age,
                application_number=a.application_number
            )
            json_res.append(json_obj)

        return HttpResponse(simplejson.dumps(json_res), content_type='application/json')
    else:
        dump = simplejson.dumps("Applicant does not exist....")
        return HttpResponse(dump, content_type='application/json')


def applicantInfo(request, admission_number):
    applicant = Applicantmodel.objects.filter(application_number=admission_number).filter(admited=1).get()

    if (applicant):
        data = {
            'name': applicant.name,
            'programme': applicant.programme_admitted.name,
            'hall': applicant.hall_admitted.hall_name,
            'fees': applicant.admission_fees,
            'firstname': applicant.first_name,
            'student': applicant.id,
            'lastname': applicant.last_name,
            'pcode': applicant.programme_admitted.code,
            'ptype': applicant.programme_admitted.type,
            'age': applicant.age,
            'application_number': applicant.application_number
        }
        dump = simplejson.dumps(data)
        return HttpResponse(dump, content_type='application/json')

    else:
        dump = simplejson.dumps("Applicant admission pending.... try again later")
        return HttpResponse(dump, content_type='application/json')


def urlify(s):
    # Remove all non-word characters (everything except numbers and letters)
    s = re.sub(r"[^\w\s]", '', s)

    # Replace all runs of whitespace with a single dash
    s = re.sub(r"\s+", '-', s)

    return s


# @login_required
def remoteSale(request, form_type, unit, location, phone, name, person, token):
    # form type refers to the form category the applicant is buying
    # unit refers to the entity selling the form eg. post office, zenith bank
    auth = ["12Y793asdy", "19FGHH9tvc7"]
    year = str(date.today().year)
    date_sold = date.today()
    # year = "2020"
    form_type = urlify(form_type)
    unit = urlify(unit)
    location = urlify(location)
    phone = urlify(phone)
    name = urlify(name)
    person = urlify(person)
    prices = {"HND": "150", "DEGREE": "150", "BTECH": "150", "MTECH": "200", "ACCESS": "150", "MATURE": "150",
              "CERTIFICATE": "150",
              "BRIDGING": "120", "DIPLOMA": "150"}
    amount = prices.get(form_type)
    if token in auth:
        form = Aspnetusers.objects.filter(soldby=unit).filter(type=form_type).filter(sold=0).filter(
            year=year).first()

        if form:

            # update system after sales
            Aspnetusers.objects.filter(soldby=unit).filter(pin=form.pin).update(sold=1, phonenumber=phone,
                                                                                branch=location, fullname=name,
                                                                                teller=person)
            sale = Sale()
            sale.pin = form.pin
            sale.serial = form.username
            sale.year = year
            sale.deleted = 0
            sale.free = 0
            sale.belongs = unit
            sale.customer_name = name
            sale.customer_phone = phone
            sale.amount = amount
            sale.date_sold = date_sold
            sale.form_type = form_type
            sale.save()
            response = {'responseCode': "01", 'responseMessage': "Success.", 'Serial': form.username,
                        'Pin': form.pin}
            return JsonResponse(response, content_type='application/json')

        else:
            # dump = simplejson.dumps("Voucher not available")
            response = {'responseCode': "09", 'responseMessage': "Failed voucher type not available.", 'Serial': "",
                        'Pin': ""}
            return JsonResponse(response, content_type='application/json')
    else:

        response = {'responseCode': "01", 'responseMessage': "Auth failed."}
        return JsonResponse(response, content_type='application/json')


def saleStatistics(request):
    query = Aspnetusers.objects.values('soldby').filter(formno__isnull=False).annotate(Count('sold')).filter(
        year="2020").filter(sold=1).filter(~Q(soldby='TPCONNECT')).filter(~Q(soldby='SMS')).filter(~Q(soldby='TEST'))

    if (query):

        return query


    else:

        return 1


@csrf_exempt
def form_sales(request):
    auth = ["12Y793asdy", "19FGHH9tvc7"]
    year = str(date.today().year)
    form_type = request.POST.get('form_type')
    unit = str.upper(request.POST.get('unit'))
    print(request.POST.get('unit'))
    location = request.POST.get('location')
    phone = request.POST.get('phone')
    name = request.POST.get('name')
    person = request.POST.get('person')
    token = request.POST.get('token')
    date_sold = date.today()
    prices = {"HND": "150", "DEGREE": "150", "BTECH": "150", "MTECH": "200", "ACCESS": "150", "MATURE": "150",
              "CERTIFICATE": "150",
              "BRIDGING": "120", "DIPLOMA": "150"}
    amount = prices.get(form_type)
    if token in auth:
        form = Aspnetusers.objects.filter(soldby=unit).filter(type=form_type).filter(sold=0).filter(
            year=year).first()

        if form:

            # update system after sales
            Aspnetusers.objects.filter(soldby=unit).filter(pin=form.pin).update(sold=1, phonenumber=phone,
                                                                                branch=location, fullname=name,
                                                                                teller=person)
            sale = Sale()
            sale.pin = form.pin
            sale.serial = form.username
            sale.year = year
            sale.deleted = 0
            sale.free = 0
            sale.belongs = unit
            sale.customer_name = name
            sale.customer_phone = phone
            sale.amount = amount
            sale.date_sold = date_sold
            sale.form_type = form_type
            sale.save()
            response = {'responseCode': "01", 'responseMessage': "Success.", 'Serial': form.username,
                        'Pin': form.pin}
            return JsonResponse(response, content_type='application/json')

        else:
            # dump = simplejson.dumps("Voucher not available")
            response = {'responseCode': "09", 'responseMessage': "Failed voucher type not available.", 'Serial': "",
                        'Pin': ""}
            return JsonResponse(response, content_type='application/json')
    else:

        response = {'responseCode': "01", 'responseMessage': "Auth failed."}
        return JsonResponse(response, content_type='application/json')


@login_required
def home(request):
    current_user = request.user.username
    setting = Configurations.objects.order_by('-id')[0]
    year = setting.working_year

    admitted = Applicantmodel.objects.filter(admitted=1).filter(yearofadmission=year).count()
    total = Applicantmodel.objects.filter(yearofadmission=year).count()
    pending = total - admitted

    admittedPercent = round(decimal.Decimal(admitted / total * 100), 2)
    pendingPercent = round(decimal.Decimal(pending / total * 100), 2)

    todayApplicants = Applicantmodel.objects.filter(yearofadmission=year).order_by('-id').all()[:10]

    # hndApplicants = Applicantmodel.objects.filter(form_type='hnd').count()
    # matureApplicants = Applicantmodel.objects.filter(form_type='mature').count()
    # btechApplicants = Applicantmodel.objects.filter(form_type='btech').count()
    # mtechApplicants = Applicantmodel.objects.filter(form_type='mtech').count()
    # diplomaApplicants = Applicantmodel.objects.filter(form_type='diploma').count()
    hndApplicants = 1
    matureApplicants = 1
    btechApplicants = 1
    mtechApplicants = 1
    diplomaApplicants = 1
    sales = saleStatistics(request)

    context = {'admitted': admitted, 'total': total, 'pending': pending, 'admittedPercent': admittedPercent,
               'pendingPercent': pendingPercent,
               'today': todayApplicants, 'hnd': hndApplicants, 'mature': matureApplicants, 'btech': btechApplicants,
               'diploma': diplomaApplicants, 'mtech': mtechApplicants
               }
    # if current_user == "finance":
    #     return redirect('applicants_index')
    # else:
    return render(request, 'dashboard/index.html', context)
    # return render(request, 'dashboard/index.html', context)


# sample api hiting with django

# def home(request):
#     if request.method == "POST":
#         apiKey=1
#         payload = {'apikey': apiKey,
#                    'zip': request.POST.get('zipcode'), }
#         response = requests.get("http://congress.api.sunlightfoundation.com/legislators/locate?callback=?",
#                                 params=payload)

def firesms(phone, message, sender, applicant, request):
    ClientID = "ifrzlixd"
    ClientSecret = "zrydysvw"

    sender = "TTU"
    # parameters to send SMS

    params = {"ClientID": ClientID, "To": phone, "Content": message, "From": sender, "ClientSecret": ClientSecret}

    r = requests.get("https://smsc.hubtel.com/v1/messages/send?", params=params)
    print(r.content)
    # url="https://api.hubtel.com/v1/messages/send?"+urllib.parse.urlencode(params)

    # urllib.request.urlopen(url)

    year = str(date.today().year) + "/" + str(date.today().year + 1)

    status = r.status_code
    data = Applicantmodel.objects.filter(applicationnumber=applicant).get()
    sms = Smsmodel()
    sms.type = "admissions notifications"
    sms.sentby = request.user
    sms.status = status
    sms.phone = phone
    sms.applicant = data.id
    sms.datesent = timezone.now()
    sms.recipient = data.applicationnumber
    sms.year = year
    sms.message = message
    sms.save()
    Applicantmodel.objects.filter(applicationnumber=applicant).update(smssent=True)


def fireCustomSMS(request):
    # $newstring = str_replace("]", "", "$message");
    # $finalstring = str_replace("[", "$", "$newstring");
    # eval("\$finalstring =\"$finalstring\" ;");

    if request.method == "POST":
        print(request.POST.get('phone'))

        message = request.POST.get('message')
        newstring = message.replace("]", "")
        finalMessage = newstring.replace("[", "")

        phones = request.POST.get('phone').split(',')

        print(phones)

        for phone in phones:
            data = Applicantmodel.objects.filter(phone=phone)[:1].get()

            application_number = data.application_number
            first_name = data.firstname
            last_name = data.lastname
            name = data.firstname
            # programme_admitted=data.programme_admitted.name
            admission_fees = str(data.admission_fees)
            mess = eval(finalMessage)

            firesms(phone, mess, request.user, data, request)

        responseData = {'status': "success", 'message': "Messages sent to applicants successfully"}

        # return redirect('applicants_index')

        return JsonResponse(responseData)


def getVoucher(formno):
    try:
        data = Aspnetusers.objects.filter(formno=formno).get()

        pin = data.pin
        serial = data.username

        info = "Pin = " + pin + " and  Serial = " + serial

        return info
    except:
        pass


@login_required
def applicants_admitted(request):
    programme = Programmemodel.objects.filter(runing=1).order_by('name').all()
    hall = Hallmodel.objects.all()
    regions = Regionmodel.objects.all()
    college = Facultymodel.objects.all()
    department = Departmentmodel.objects.all()
    faculty = Facultymodel.objects.all()
    countries = Countrymodel.objects.all()

    data = Applicantmodel.objects.filter(admited='1').all()

    if ('gender' in request.GET.keys() and request.GET.get('gender')):
        data = Applicantmodel.objects.all().filter(admited=1).filter(gender=request.GET.get('gender'))

    if ('search' in request.GET.keys() and request.GET.get('search')):
        data = Applicantmodel.objects.all().filter(Q(applicationnumber=request.GET.get('search')) |
                                                   Q(name__icontains=request.GET.get('search')))

    applicants = data.order_by('-id')
    total = applicants.count()

    # Provide Paginator with the request object for complete querystring generation

    paginator = Paginator(applicants, 150)  # Show 150 contacts per page

    page = request.GET.get('page')
    try:
        applicants = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        applicants = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        applicants = paginator.page(paginator.num_pages)

    return render(request, 'old/applicants/admitted.html',
                  {'data': applicants, 'total': total, 'hall': hall, 'programme': programme, 'regions': regions
                      , 'college': college, 'department': department, 'faculty': faculty, 'countries': countries,
                   })


@login_required
def applicants_index(request):
    setting = Configurations.objects.order_by('-id')[0]
    start_admit = setting.start_admit
    print(start_admit)
    programme = Programmemodel.objects.order_by('name').all()
    hall = Hallmodel.objects.all()
    regions = Regionmodel.objects.all()
    college = Facultymodel.objects.all()
    department = Departmentmodel.objects.all()
    faculty = Facultymodel.objects.all()
    countries = Countrymodel.objects.all()
    school = Formerschoolmodel.objects.all()

    # nextYear = datetime.date.today().year + 1
    # year = str(datetime.date.today().year)+str("/")+str(nextYear)

    records = Applicantmodel.objects.get_queryset()
    gender = request.GET.get('gender', None)
    status = request.GET.get('status', None)
    year = request.GET.get('year', None)
    program = request.GET.get('program', None)
    search = request.GET.get('search', None)
    departments = request.GET.get('department', None)
    nationality = request.GET.get('nationality', None)
    former = request.GET.get('school', None)
    region = request.GET.get('region', None)
    halls = request.GET.get('hall', None)
    form_type = request.GET.get('type', None)
    admission_type = request.GET.get('admission_type', None)
    referral = request.GET.get('referral', None)
    faculties = request.GET.get('faculty', None)
    sorts = request.GET.get('sort', None)
    limit = request.GET.get('limit', None)
    paid = request.GET.get('paid', None)
    owing = request.GET.get('owing', None)
    sms = request.GET.get('sms', None)
    programtype = request.GET.get('programtype', None)
    limit_range = 100

    if gender:
        records = records.filter(gender=gender)
        # print(str(records.query))

    if status:
        records = records.filter(status=status)
        # print(str(records.query))

    if program:
        records = records.filter(programmeadmittedid=program)
        # print("program")

    if former:
        records = records.filter(formerschoolnewid=former)

    if region:
        records = records.filter(regionid=region)

    if paid == "paid":
        records = records.filter(feespaid__gt=0)
        # print(str(records.query))
    if paid == "owing":
        records = records.filter(feespaid=0)

    if admission_type:
        records = records.filter(admissiontype=admission_type)
    if referral:
        records = records.filter(referrals=referral)

    if year:
        records = records.filter(yearofadmission=year)

    if departments:
        records = records.filter(programmeadmittedid__department=departments)
    # programme types
    if programtype:
        records = records.filter(programmeadmittedid__type=programtype)

    if faculties:
        records = records.filter(programmeadmittedid__department__faculty=faculties)

    if form_type:
        records = records.filter(applicationuserid__type=form_type)

    if sms:
        records = records.filter(smssent=sms)

    if halls:
        records = records.filter(halladmitted=halls)

    if nationality:
        if nationality == "58":
            records = records.filter(nationalityid=58)
        elif nationality == "0":
            records = records.filter(~Q(nationalityid=58))
            # print(str(records.query))

    if search:
        if search.isnumeric():
            records = records.filter(Q(applicationnumber=search))
        else:
            records = records.filter(
                Q(firstname__icontains=search) |
                Q(lastname__icontains=search)

                |
                Q(applicationuserid__pin=search)

                |
                Q(applicationuserid__username=search)

            )

    if limit:
        limit_range = limit

    if sorts:

        applicants = records
        total = applicants.count()
        paginator = Paginator(applicants.order_by(str(sorts)), limit_range)
        print(str(paginator))

    else:
        applicants = records
        total = applicants.count()

        paginator = Paginator(applicants.order_by('-id').order_by('lastname').order_by('-applicationnumber'),
                              limit_range)

    try:
        page = int(request.GET.get('page', '1'))
    except:
        page = 1

    try:
        applicants = paginator.page(page)

    except(EmptyPage, InvalidPage):
        applicants = paginator.page(paginator.num_pages)
    sessionData = []
    for app in applicants:
        json_obj = dict({'firstname': app.firstname, 'lastname': app.lastname,
                         'gender': app.gender, 'phone': app.phone,
                         'applicationnumber': app.applicationnumber,

                         'name': app.firstname, 'admission_type': app.admissiontype

                         })

        sessionData.append(json_obj)

    request.session['applicants'] = sessionData
    return render(request, 'applicants/applicants.html',
                  {'data': applicants,
                   'total': total, 'hall': hall, 'programme': programme, 'regions': regions
                      , 'college': college, 'department': department, 'faculty': faculty, 'countries': countries,
                   'school': school, 'start_admit': start_admit
                   })


@login_required
def sell(request):
    if request.method == "GET":
        return render(request, "sale_of_forms/sell.html")
    else:
        prices = {"HND": "150", "DEGREE": "150", "BTECH": "150", "MTECH": "200", "ACCESS": "150", "MATURE": "150",
                  "CERTIFICATE": "150",
                  "BRIDGING": "120", "DIPLOMA": "150"}
        if request.POST.get('name') and request.POST.get('phone') and request.POST.get('type'):
            name = request.POST.get('name')
            phone = request.POST.get('phone')
            type = request.POST.get('type')
            year = str(date.today().year)
            current_user = request.user.id
            if request.user.username == "tpconnect":
                free = 1
            else:
                free = 0

            if request.user.username == 'tpconnect':
                voucher = Aspnetusers.objects.filter(sold=0).filter(soldby="TPCONNECT").filter(type=type).filter(
                    year=year).first()
            else:

                voucher = Aspnetusers.objects.filter(sold=0).filter(soldby="TTU").filter(type=type).filter(
                    year=year).first()

            # print(requests)

            sale = Sale()
            sale.form_type = type
            sale.sold_by = request.user
            sale.date_sold = str(date.today())
            sale.customer_name = name
            sale.customer_phone = phone
            sale.pin = voucher.pin
            sale.serial = voucher.username
            sale.amount = prices.get(type)
            sale.belongs = "TTU"
            sale.free = free
            sale.year = year
            sale.deleted = 0

            sale.save()

            Aspnetusers.objects.filter(pin=voucher.pin).update(sold=1, fullname=name, phonenumber=phone)

            records = Sale.objects.filter(free=0).filter(year=str(date.today().year)).filter(deleted=0).filter(
                sold_by_id=current_user).all().order_by("-id")

            return redirect("print_receipt", pin=voucher.pin)


def sales(request):
    total = 0.0

    limit = request.GET.get('limit', None)
    type = request.GET.get('type', None)
    year = request.GET.get('year', None)

    search = request.GET.get('search', None)
    to = request.GET.get('to', None)
    froms = request.GET.get('from', None)

    sold = request.GET.get('sold', None)

    limit_range = 100

    current_user = request.user.id

    records = Sale.objects.filter(free=0).filter(year=str(date.today().year)).filter(deleted=0).filter(
        ~Q(belongs='WEBMASTER')).filter(
        ~Q(belongs='TPCONNECT')).filter(
        ~Q(customer_name='Gad')).all().order_by("-id")

    if type:
        records = records.filter(form_type=type)

    if sold:
        records = records.filter(Q(belongs=sold) |
                                 Q(belongs=sold))

    if to and froms:
        end = datetime.strptime(to, '%Y-%m-%d')
        start = datetime.strptime(froms, '%Y-%m-%d')
        records = Sale.objects.filter(date_sold__range=(start, end)).all()

        print(records)
    if froms:
        start = datetime.strptime(froms, '%Y-%m-%d')
        records = Sale.objects.filter(date_sold=(start)).all()

        print(records)
    if year:
        records = records.filter(year=year)

    if search:
        records = records.filter(Q(customer_name=search) |
                                 Q(customer_name__icontains=search) | Q(customer_phone=search))

    if limit:
        limit_range = limit

    data = records

    total = records.count()

    paginator = Paginator(data.order_by('-id'), limit_range)

    try:
        page = int(request.GET.get('page', '1'))
    except:
        page = 1

    try:
        data = paginator.page(page)
    except(EmptyPage, InvalidPage):
        data = paginator.page(paginator.num_pages)

    return render(request, "sale_of_forms/sales.html", context={'data': data, 'total': total})


@login_required
def print_receipt(request, pin):
    data = Sale.objects.filter(pin=pin).get()

    return render(request, "sale_of_forms/receipt.html", context={'data': data})


@login_required
def delete_sales(request):
    id = request.GET.get('id')
    Sale.objects.filter(id=id).update(deleted=1)
    responseData = {'status': "success", 'message': "Sales deleted successfully"}
    return JsonResponse(responseData)


@login_required
def update_programme(request):
    id = request.POST.get('id')

    running = request.POST.get('running')

    admitting = str.title(request.POST.get('admitting'))

    running = str.title(running)
    query = Programmemodel.objects.filter(id=id).update(runing=running, admitting=admitting)
    if query:
        responseData = {'status': "success", 'code': 1}

    else:
        responseData = {'status': "error", 'code': 0}

    return JsonResponse(responseData)


@login_required
def programmes(request):
    data = Programmemodel.objects.order_by('type').all()

    total = data.count()

    paginator = Paginator(data.order_by('type'), 50)

    try:
        page = int(request.GET.get('page', '1'))
    except:
        page = 1

    try:
        data = paginator.page(page)
    except(EmptyPage, InvalidPage):
        data = paginator.page(paginator.num_pages)

    return render(request, "programmes/index.html", context={'data': data, 'total': total})


@login_required
def upload_photo(request):
    if request.method == 'GET':
        return render(request, 'upload/photo.html')
    else:

        media_root = settings.MEDIA_ROOT
        # media_root = ""
        host = '45.33.4.164'
        username = 'tpconnect'
        port = 44444
        password = 'PRINT45dull'
        server_path = '/var/www/html/photos/public/albums/thumbnails'

        form = request.POST.get("forms")
        myfile = request.FILES['file']

        check = Applicantmodel.objects.filter(applicationnumber=form).get()

        if not check.applicationnumber:

            messages.error(request, "Applicant with " + form + "does not exist!")
            return redirect("upload_photo")
        else:

            valid_extensions = [".jpg", ".JPG", ".JPEG", ".jpeg"]

            ext = os.path.splitext(str(myfile))[1]

            if ext in valid_extensions and myfile.size <= 250000:
                filename = "%s%s" % (form, ext)

                # print(filename)
                fs = FileSystemStorage()

                filename = fs.save(filename, myfile)
                uploaded_file_url = fs.url(filename)
                url = media_root + "/" + filename

                a = sftp_upload_file(host, port, username, password, server_path, url)
                print("uploaded.. " + str(a))
                messages.success(request, 'Photo successfully uploaded!')
                return redirect("upload_photo")
            else:
                messages.error(request, 'Error upload photo!')
                return redirect("upload_photo")


def sftp_upload_file(host, port, user, password, server_path, local_path, timeout=10):
    try:
        # create transport object.
        cnopts = pysftp.CnOpts()
        cnopts.hostkeys = None
        with pysftp.Connection(host=host, username=user, password=password, port=port, private_key=".ppk",
                               cnopts=cnopts) as sftp:
            print("Connection succesfully stablished ... ")

            # Define the file that you want to upload from your local directorty
            # or absolute "C:\Users\sdkca\Desktop\TUTORIAL2.txt"
            localFilePath = local_path

            # Define the remote path where the file will be uploaded
            remoteFilePath = server_path
            sftp.chdir("/var/www/html/photos/public/albums/thumbnails")
            sftp.put(localFilePath)

            sftp.close()
        return True
    except Exception as e:
        print(e)
        return False


# live search using ajax consuming json response
def applicant_live_search(request):
    term = request.GET.get('term')

    results = []

    data = Aspnetusers.objects.filter(
        Q(username__icontains=term) | Q(pin__icontains=term) | Q(
            phonenumber__icontains=term) |
        Q(fullname__icontains=term) | Q(formno=term))[
           :100].all()
    print(data.__len__())
    for row in data:
        if row.formno:
            results.append(dict([('id', row.id), ('value', str(row.formno) + ',' + row.username)]))
    print(data.query)
    return JsonResponse(results, safe=False)


def block_unblock(request):
    if request.method == 'GET':
        return render(request, 'support/block_unblock.html')
    else:
        applicant = request.POST.get('q')
        a, b = applicant.split(',')

        data = Applicantmodel.objects.filter(applicationnumber=a).get()

        print(data)

        return render(request, 'support/applicant_details.html', context={'applicant': data})


def modify_applicant_form(request):
    applicant = request.POST.get('applicant')
    action = request.POST.get('action')
    name = request.POST.get('name')

    if action == "1":
        Aspnetusers.objects.filter(formno=applicant).update(finalized=1)
        responseData = {'status': "success", 'message': name + " form blocked successfully"}
        return JsonResponse(responseData, safe=False)
    else:
        Aspnetusers.objects.filter(formno=applicant).update(finalized=0)
        responseData = {'status': "success", 'message': name + " form unblocked successfully"}

        return JsonResponse(responseData, safe=False)


def changeForm(request):
    if request.method == 'GET':
        return render(request, 'support/voucher.html')
    else:
        applicant = request.POST.get('q')
        a, b = applicant.split(',')

        data = Aspnetusers.objects.filter(formno=a).get()

        # print(data.firstname)

        return render(request, 'support/voucher_details.html', context={'applicant': data})


def changeVoucher(request):
    applicant = request.POST.get('applicant')
    action = request.POST.get('type')

    Aspnetusers.objects.filter(formno=applicant).update(type=action, formcompleted=0, started=0, finalized=0)

    # emp = Applicantmodel.objects.get(applicationnumber=applicant)

    # emp.delete()

    responseData = {'status': "success", 'message': " form changed to " + action + " successfully"}
    return JsonResponse(responseData, safe=False)


def createSettings(request):
    if request.method == "POST":

        if request.POST.get('year') and request.POST.get('name') and request.POST.get(
                'medicals_starts') and request.POST.get('medicals_ends') and request.POST.get(
            'deadline') and request.POST.get('reporting') and request.POST.get('orient_ends') and request.POST.get(
            'orient_starts'):

            print("gad")

            year = request.POST.get('year')
            officer = request.POST.get('name')
            medicals_start = request.POST.get('medicals_starts')
            medicals_end = request.POST.get('medicals_ends')
            deadline = request.POST.get('deadline')
            reporting = request.POST.get('reporting')
            orient_ends = request.POST.get('orient_ends')
            orient_starts = request.POST.get('orient_starts')

            Configurations.objects.update_or_create(working_year=year, defaults={"officer": officer,
                                                                                 "medical_examinations_starts": medicals_start,
                                                                                 "medical_examinations_ends": medicals_end,
                                                                                 "fees_deadline": deadline,
                                                                                 "orientation_starts": orient_starts,
                                                                                 "orientation_ends": orient_ends,
                                                                                 "reporting": reporting,
                                                                                 "start_admit": False,
                                                                                 "running": '0',
                                                                                 "created_at": date.today(),
                                                                                 "updated_at": date.today(),
                                                                                 "created_by_id": request.user.id}

                                                    )
            # config.save()

            responseData = {'status': "success", 'message': "System settings updated successfully"}

            return JsonResponse(responseData)

        else:
            responseData = {'status': "error", 'message': "Please fill all the required fields"}

            return JsonResponse(responseData)


@login_required
def mainConfiguration(request):
    if request.method == 'GET':
        data = Configurations.objects.all().order_by("-id")

        year = ["2018/2019", "2019/2020", "2020/2021", "2021/2022", "2022/2023"]

        context = {"data": data, "years": year}

        return render(request, 'settings/calender.html', context)


@login_required
def deleteSetting(request):
    trial = Configurations.objects.get(pk=request.GET.get('id')).delete()

    print(trial)

    if trial:
        responseData = {'status': "success", 'message': "System settings updated successfully"}

        return JsonResponse(responseData)
    else:
        responseData = {'status': "error", 'message': "Error deleting data"}

        return JsonResponse(responseData)


def letter(request, object_id):
    applicant = get_object_or_404(Applicantmodel, pk=object_id)

    image_logo = os.path.join(os.getcwd(), 'admissions/static/assets/img/', 'header' + '.jpg')

    print(image_logo)
    # setting = Configurations.objects.order_by('-id')[0]
    setting = Configurations.objects.filter(working_year=applicant.yearofadmission).first()
    print("start" + setting.working_year)
    current_year = setting.working_year
    context = {'applicant': applicant, 'logo': image_logo, 'year': setting}

    print(applicant.admissiontype)

    template = get_template('letters/2020/regular.html')
    pdf = ""
    html = template.render(context)
    # current_year=str(date.today().year) + "/"+ str((date.today().year)+1)

    if applicant.yearofadmission == '2017/2018':
        if applicant.admissiontype == 'regular' or applicant.admissiontype == 'mature':

            return render(request, 'letters/2017/regular.html', context)


        elif applicant.admissiontype == 'technical':

            return render(request, 'letters/2017/technicalLetter.html', context)


        elif applicant.admissiontype == 'diploma_education':

            pdf = render_to_pdf('letters/2017/dipEdu_letter.html', context)


        elif applicant.admissiontype == 'conditional':

            return render(request, 'letters/2017/conditional.html', context)


        elif applicant.admissiontype == 'bridging':

            return render(request, 'letters/2017/bridging.html', context)

        elif applicant.admissiontype == 'foreign':

            return render(request, 'letters/2017/regular.html', context)

        else:

            return render(request, 'letters/2017/regular.html', context)

    if applicant.yearofadmission == '2018/2019':
        if applicant.admissiontype == 'regular' or applicant.admissiontype == 'mature':

            return render(request, 'letters/2018/regular.html', context)


        elif applicant.admissiontype == 'technical':

            return render(request, 'letters/2018/technicalLetter.html', context)


        elif applicant.admissiontype == 'diploma_education':

            pdf = render_to_pdf('letters/2018/dipEdu_letter.html', context)


        elif applicant.admissiontype == 'conditional':

            return render(request, 'letters/2018/conditional.html', context)


        elif applicant.admissiontype == 'bridging':

            return render(request, 'letters/2018/bridging.html', context)

        elif applicant.admissiontype == 'foreign':

            return render(request, 'letters/2018/regular.html', context)

        else:

            return render(request, 'letters/2018/regular.html', context)

    if applicant.yearofadmission == '2017/2018':
        if applicant.admissiontype == 'regular' or applicant.admissiontype == 'mature':

            return render(request, 'letters/2017/regular.html', context)


        elif applicant.admissiontype == 'technical':

            return render(request, 'letters/2017/technicalLetter.html', context)


        elif applicant.admissiontype == 'diploma_education':

            pdf = render_to_pdf('letters/2017/education.html', context)


        elif applicant.admissiontype == 'conditional':

            return render(request, 'letters/2017/conditional.html', context)


        elif applicant.admissiontype == 'bridging':

            return render(request, 'letters/2017/bridging.html', context)

        elif applicant.admissiontype == 'foreign':

            return render(request, 'letters/2017/regular.html', context)

        else:

            return render(request, 'letters/2017/regular.html', context)


    elif applicant.yearofadmission == '2019/2020':

        if applicant.admissiontype == 'reqular' or applicant.admissiontype == 'mature':

            pdf = render_to_pdf('letters/2019/regular.html', context)

        elif applicant.admissiontype == 'technical':

            pdf = render_to_pdf('letters/2019/technicalLetter.html', context)

        elif applicant.admissiontype == 'diploma_education':

            pdf = render_to_pdf('letters/2019/dipEdu_letter.html', context)


        elif applicant.admissiontype == 'conditional':

            pdf = render_to_pdf('letters/2019/conditional.html', context)

        elif applicant.admissiontype == 'bridging':
            pdf = render_to_pdf('letters/2019/bridging.html', context)


        elif applicant.admissiontype == 'foreign':
            pdf = render_to_pdf('letters/2019/international.html', context)

        else:

            pdf = render_to_pdf('letters/2019/regular.html', context)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "letter_%s.pdf" % ("12341231")
            content = "inline; filename='%s'" % (filename)
            download = request.GET.get("download")
            if download:
                content = "attachment; filename='%s'" % (filename)
            response['Content-Disposition'] = content

            return response

    elif applicant.yearofadmission == '2020/2021':
        print("year of admissions is " + applicant.yearofadmission)
        if applicant.admissiontype == 'reqular' or applicant.admissiontype == 'mature':

            pdf = render_to_pdf('letters/2020/regular.html', context)

        elif applicant.admissiontype == 'technical':

            pdf = render_to_pdf('letters/2020/technicalLetter.html', context)

        elif applicant.admissiontype == 'diploma_education':

            pdf = render_to_pdf('letters/2020/dipEdu_letter.html', context)


        elif applicant.admissiontype == 'conditional':

            pdf = render_to_pdf('letters/2020/conditional.html', context)

        elif applicant.admissiontype == 'bridging':
            pdf = render_to_pdf('letters/2020/bridging.html', context)


        elif applicant.admissiontype == 'foreign':
            pdf = render_to_pdf('letters/2020/international.html', context)

        else:

            pdf = render_to_pdf('letters/2020/regular.html', context)

        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "letter_%s.pdf" % ("12341231")
            content = "inline; filename='%s'" % (filename)
            download = request.GET.get("download")
            if download:
                content = "attachment; filename='%s'" % (filename)
            response['Content-Disposition'] = content

            return response

    elif applicant.yearofadmission == current_year:
        # path="letters/"+str(date.today().year)+"/"
        path = "letters/2021/"

        if applicant.admissiontype == 'reqular' or applicant.admissiontype == 'mature' or applicant.admissiontype == 'distance':

            pdf = render_to_pdf(path + "regular.html", context)

        elif applicant.admissiontype == 'technical':

            pdf = render_to_pdf(path + "technicalLetter.html", context)

        elif applicant.admissiontype == 'diploma_education':

            pdf = render_to_pdf(path + "education.html", context)


        elif applicant.admissiontype == 'distance':

            pdf = render_to_pdf(path + "distance.html", context)
        elif applicant.admissiontype == 'proficiency':

            pdf = render_to_pdf(path + "proficiency.html", context)

        elif applicant.admissiontype == 'bridging':
            pdf = render_to_pdf(path + "bridging.html", context)


        elif applicant.admissiontype == 'foreign':
            pdf = render_to_pdf(path + "international.html", context)

        else:

            pdf = render_to_pdf(path + "regular.html", context)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "letter_%s.pdf" % ("12341231")
            content = "inline; filename='%s'" % (filename)
            download = request.GET.get("download")
            if download:
                content = "attachment; filename='%s'" % (filename)
            response['Content-Disposition'] = content

            return response

    return HttpResponse("PDF downloaded")


@login_required
def openadmissions(request):
    id = request.GET.get('id')

    admit = request.GET.get('admit')
    if admit == "1":
        trial = Configurations.objects.filter(id=id).update(start_admit=True)
    else:
        trial = Configurations.objects.filter(id=id).update(start_admit=False)

    if trial:
        responseData = {'status': "success", 'message': "System opened for admissions successfully"}

        return JsonResponse(responseData)
    else:
        responseData = {'status': "error", 'message': "Error opening system"}

        return JsonResponse(responseData)


@login_required
def openapplication(request):
    id = request.GET.get('id')

    application = request.GET.get('application')
    print(application)
    if application == "1":
        trial = Configurations.objects.filter(id=id).update(running="1")
    else:
        trial = Configurations.objects.filter(id=id).update(running="0")

    print(trial)

    if trial:
        responseData = {'status': "success", 'message': "Application opened for applicants successfully"}

        return JsonResponse(responseData)
    else:
        responseData = {'status': "error", 'message': "Error opening applications"}

        return JsonResponse(responseData)


@login_required
def messages_sms(request):
    records = Smsmodel.objects.get_queryset()

    limit = request.GET.get('limit', None)
    status = request.GET.get('status', None)

    search = request.GET.get('search', None)
    year = request.GET.get('year', None)
    limit_range = 100
    x = slice(4)
    if year:
        records = records.filter(datesent__year=year[x])

    if status:
        records = records.filter(status=status)

    if search:
        records = records.filter(recipient=search)

    if limit:
        limit_range = limit

    data = records

    total = records.count()

    paginator = Paginator(data.order_by('-id'), limit_range)

    try:
        page = int(request.GET.get('page', '1'))
    except:
        page = 1

    try:
        data = paginator.page(page)
    except(EmptyPage, InvalidPage):
        data = paginator.page(paginator.num_pages)

    context = {'data': data, 'total': total}
    return render(request, 'sms/messages.html', context)


@login_required
def sms_admissions(request):
    return render(request, 'sms/send_message.html')


def getVoucher(formno):
    try:
        data = Aspnetusers.objects.filter(formno=formno).get()

        pin = data.pin
        serial = data.username

        info = "Pin = " + pin + " and  Serial = " + serial

        return info
    except:
        pass


def send_notifications(request):
    return None;


@login_required
def sendBulkSMS(request):
    applicants = request.session['applicants']
    # print(applicants)
    message = ""
    sender = request.user
    for applicant in applicants:

        formDetails = str(getVoucher(applicant["applicationnumber"]))

        if applicant["admission_type"] == "regular":

            message = applicant[
                          "firstname"] + "You have been admitted to TTU. Visit admissions.ttuportal.com to print your admission " \
                                         "letter." + formDetails

        elif applicant["admission_type"] == "conditional":
            message = applicant[
                          "firstname"] + "You have been admitted to TTU. Visit admissions.ttuportal.com to print your admission " \
                                         "letter." + formDetails

        elif applicant["admission_type"] == "technical":
            message = applicant[
                          "firstname"] + "You have been admitted to TTU. Visit admissions.ttuportal.com to print your admission " \
                                         "letter." + formDetails

        elif applicant["admission_type"] == "mature":
            message = applicant[
                          "firstname"] + "You have been admitted to TTU. Visit  admissions.ttuportal.com to print your " \
                                         "admission letter." + formDetails
        else:
            message = applicant[
                          "firstname"] + "You have been admitted to TTU. Visit admissions.ttuportal.com to print your admission " \
                                         "letter." + formDetails

        firesms(applicant["phone"], message, sender, applicant["applicationnumber"], request)

        print(applicant["phone"])
        applicantNum = applicant["applicationnumber"]

        # Applicantmodel.objects.filter(applicationnumber=applicantNum).update(smssent=True)

    request.session['applicants'] = None
    return redirect('applicants_index')


def sendBulkSMSNotice(request):
    applicants = request.session['applicants']
    message = request.POST.get('message', None)
    newstring = message.replace("]", "")
    finalMessage = newstring.replace("[", "")

    sender = request.user
    for applicant in applicants:
        application_number = applicant.applicationnumber
        first_name = applicant.firstname
        last_name = applicant.lastname
        name = last_name + first_name
        hall_admitted = "none"
        programme_admitted = applicant.programmeadmitted_id.name
        admission_fees = str(applicant.admissionfees)
        mess = eval(finalMessage)
        firesms(applicant.phone, mess, sender, applicant, request)

        print(mess)
    # request.session['applicants'] = None
    return HttpResponse("sms sent")
    # return redirect('applicants_index')


#
# after successfull send data to srms to create student files for applicant
@login_required
def admit(request):
    try:

        # get the academic year info for an applicant

        # responseData = {'status': "error", 'message': "Admission Ended"}
        # print("hello")
        # return redirect('applicants_index')

        # return HttpResponse(responseData)

        applicant = request.POST.get('applicant')
        hall = request.POST.get('hall')
        # section = request.POST.get('section')
        type = request.POST.get('type')
        residentialstatus = request.POST.get('resident')
        programme = request.POST.get('program')
        levelAdmitted = request.POST.get('level')

        # details of the admitted programme

        applicantRecord = Applicantmodel.objects.filter(id=applicant).first()
        # print(applicantRecord.firstname)
        setting = Configurations.objects.filter(working_year=applicantRecord.yearofadmission).first()

        # year=setting.working_year
        # get fees of admitted programme
        headers = {'Content-type': 'application/json'}
        # data = requests.get(
        #     "https://srms.ttuportal.com/api/get/fees/" + applicantRecord.programmeadmittedid.code + "/program/" + levelAdmitted + "/level/2020/2021/year",
        #     headers=headers)
        # fees=data.text

        current_user = request.user
        import pytz
        now = datetime.now()

        UTC = pytz.utc
        if applicant and hall and programme and type and residentialstatus and programme:
            checker = Applicantmodel.objects.filter(id=applicant).update(admitted=True, programmeadmittedid=programme,

                                                                         admissiontype=type,
                                                                         residentialstatus=residentialstatus,
                                                                         halladmitted=hall,
                                                                         smssent=False, status='admitted',
                                                                         dateadmitted=datetime.now(UTC),
                                                                         letterprinted=False,
                                                                         publish=1,
                                                                         admittedby=current_user.id,
                                                                         ProgrammeAdmittedId=programme,
                                                                         leveladmitted=levelAdmitted,
                                                                         lettersettings=setting.id

                                                                         )
            Aspnetusers.objects.filter(formno=applicantRecord.applicationnumber).update(admitted=True)

            returnToSrms(applicant)

            # print("checksss" + str(checker))
            if checker:
                # call sms service
                # refire_sms( request,applicantRecord.application_number)
                responseData = {'status': "success", 'message': "Applicant admitted successfully"}

                # return redirect('applicants_index')

                return JsonResponse(responseData)
            else:
                responseData = {'status': "error", 'message': "Please confirm program has fees before admitting"}

                # return redirect('applicants_index')

                return JsonResponse(responseData)



    except:

        responseData = {'status': "error", 'message': "Error admitting applicant"}

        # return redirect('applicants_index')

        return HttpResponse(responseData)


# cancel the admission of applicant
## send request to srms to delete applicant from student records

def revoke_admission(request):
    try:
        applicant = request.GET.get('applicant')

        # todayDate = datetime.now()

        current_user = request.user

        Applicantmodel.objects.filter(id=applicant).update(admitted=False, status="Applicant", publish=0)

        responseData = {'status': "success", 'message': "Applicant admission cancelled successfully"}

        # return redirect('applicants_index')

        return JsonResponse(responseData)


    except:

        responseData = {'status': "error", 'message': "Error reversing admission"}

        # return redirect('applicants_index')

        return HttpResponse(responseData)


# call this function anytime an applicant status is updated
# send the data to srms.ttuportal.com to update his data

def returnToSrms(applicant):
    applicantData = Applicantmodel.objects.filter(id=applicant).first()
    print("firstname" + applicantData.firstname)
    response_data = {'firstname': applicantData.firstname, 'lastname': applicantData.lastname,
                     'gender': applicantData.gender, 'title': applicantData.title,
                     'dob': str(applicantData.dateofbirth),
                     'name': applicantData.lastname + " " + applicantData.firstname + " " + applicantData.middlename,
                     'othernames': applicantData.middlename,
                     'finance': applicantData.sourceoffinance, 'marital': applicantData.maritalstatus,
                     'date_admitted': str(applicantData.dateadmitted.strftime("%Y-%m-%d %H:%M:%S")),
                     'level': applicantData.leveladmitted, 'stno': applicantData.applicationnumber,
                     'resident': (applicantData.residentialstatus,), 'disable': applicantData.disability,
                     'guardian_relation': applicantData.guardianrelationship,
                     'guardian_phone': applicantData.guardianphone,
                     'guardian_occupation': applicantData.guardianoccupation,
                     'religion': applicantData.religionid.name,
                     'guardian_name': applicantData.guardianname,
                     'region': applicantData.regionid.name,
                     'status': "admitted", 'fees': applicantData.guardianname, 'type': applicantData.admissiontype,
                     'program': applicantData.programmeadmittedid.code, 'country': applicantData.nationalityid.name,
                     'email': applicantData.email, 'phone': applicantData.phone, 'hometown': applicantData.hometown,
                     'section': applicantData.sectionadmitted, 'address': applicantData.address,
                     'age': applicantData.age, 'hall': applicantData.halladmitted.id
                     }
    # data.append(json_obj)
    print("hiiiii")
    dump = simplejson.dumps(response_data)

    headers = {'Content-type': 'application/json'}

    r = requests.post('https://srms.ttuportal.com/api/send/realtime', dump, headers=headers)
    print(r.text)
    print("gone")
    return redirect('applicants_index')
    # return HttpResponse(dump, content_type='application/json')


# call this function anytime an applicant status is updated
# send the data to srms.ttuportal.com to update his data
def returnToSrmsLoop(applicant):
    # applicant= request.GET.get('applicant')
    applicantData = Applicantmodel.objects.filter(id=applicant).get()
    data = []

    json_obj = dict(
        {
            'firstname': applicantData.firstname, 'lastname': applicantData.lastname,
            'gender': applicantData.gender, 'title': applicantData.title,
            'dob': str(applicantData.dateofbirth),
            'name': applicantData.lastname + applicantData.firstname + applicantData.middlename,
            'othernames': applicantData.middlename,
            'finance': applicantData.sourceoffinance, 'marital': applicantData.maritalstatus,
            'date_admitted': str(applicantData.dateadmitted.strftime("%Y-%m-%d %H:%M:%S")),
            'level': applicantData.leveladmitted, 'stno': applicantData.applicationnumber,
            'resident': (applicantData.residentialstatus,), 'disable': applicantData.disability,
            'guardian_relation': applicantData.guardianrelationship,
            'guardian_phone': applicantData.guardianphone,
            'guardian_occupation': applicantData.guardianoccupation,
            'religion': applicantData.religionid.name,
            'guardian_name': applicantData.guardianname,
            'region': applicantData.regionid.name,
            'status': "admitted", 'fees': applicantData.guardianname,
            'type': applicantData.admissiontype,
            'program': applicantData.programmeadmittedid.code,
            'country': applicantData.nationalityid.name,
            'email': applicantData.email, 'phone': applicantData.phone,
            'hometown': applicantData.hometown,
            'section': applicantData.sectionadmitted, 'address': applicantData.address,
            'age': applicantData.age, 'hall': applicantData.halladmitted

        })

    data.append(json_obj)
    dump = simplejson.dumps(data)
    print(dump)
    print(dump.__len__())
    headers = {'Content-type': 'application/json'}

    print(requests.post('https://srms.ttuportal.com/api/send', dump, headers=headers))
    # return redirect('applicants_index')

    return HttpResponse("Hi")


def sendToSrms(request):
    setting = Configurations.objects.order_by('-id')[0]
    year = setting.working_year
    applicantDatas = Applicantmodel.objects.filter(admitted=True).filter(
        status="admitted").filter(publish=0).filter(yearofadmission=year).all()
    data = []
    for applicantData in applicantDatas.all():
        name = applicantData.lastname + applicantData.firstname + applicantData.middlename
        json_obj = dict(
            {
                'firstname': applicantData.firstname, 'lastname': applicantData.lastname,
                'gender': applicantData.gender, 'title': applicantData.title,
                'dob': str(applicantData.dateofbirth),
                'name': applicantData.lastname + applicantData.firstname + applicantData.middlename,
                'othernames': applicantData.middlename,
                'finance': applicantData.sourceoffinance, 'marital': applicantData.maritalstatus,
                'date_admitted': str(applicantData.dateadmitted.strftime("%Y-%m-%d %H:%M:%S")),
                'level': applicantData.leveladmitted, 'stno': applicantData.applicationnumber,
                'resident': (applicantData.residentialstatus,), 'disable': applicantData.disability,
                'guardian_relation': applicantData.guardianrelationship,
                'guardian_phone': applicantData.guardianphone,
                'guardian_occupation': applicantData.guardianoccupation,
                'religion': applicantData.religionid.name,
                'guardian_name': applicantData.guardianname,
                'region': applicantData.regionid.name,
                'status': "admitted", 'fees': applicantData.guardianname,
                'type': applicantData.admissiontype,
                'program': applicantData.programmeadmittedid.code,
                'country': applicantData.nationalityid.name,
                'email': applicantData.email, 'phone': applicantData.phone,
                'hometown': applicantData.hometown,
                'section': applicantData.sectionadmitted, 'address': applicantData.address,
                'age': applicantData.age, 'hall': applicantData.halladmitted

            })

        data.append(json_obj)
        Applicantmodel.objects.filter(applicationnumber=applicantData.applicationnumber).update(publish=1)

    dump = simplejson.dumps(data)
    print(dump)
    print(dump.__len__())
    headers = {'Content-type': 'application/json'}

    print(requests.post('https://srms.ttuportal.com/api/send', dump, headers=headers))
    # return redirect('applicants_index')
    # return redirect('applicants_index')
    return HttpResponse(dump, content_type='application/json')


def getFees(programme):
    return 0


@login_required
def generateOutreach(request):
    if request.method == "GET":
        return render(request, "sale_of_forms/outreach.html")
    else:

        setting = Configurations.objects.order_by('-id')[0]
        year = setting.working_year[0:4]
        prices = {"HND": "150", "BTECH": "150", "MTECH": "200", "ACCESS": "150", "MATURE": "150", "CERTIFICATE": "150",
                  "BRIDGING": "150", "DIPLOMA": "150"}
        if request.POST.get('number') and request.POST.get('type'):

            number = int(request.POST.get('number'))
            type = request.POST.get('type')

            if request.user.username == "webmaster" or request.user.username == "tpconnect":
                free = 1
            else:
                free = 0

            voucher = Aspnetusers.objects.filter(sold=0).filter(soldby="OUTREACH").filter(type=type).filter(
                year=year).all()[:number]

            # print(requests)

            print("total is " + str(voucher.count()) + year)
            for data in voucher:
                sale = Sale()
                sale.form_type = type
                sale.sold_by = request.user
                sale.date_sold = str(date.today())
                sale.customer_name = "OUTREACH"
                sale.customer_phone = "000000000"
                sale.pin = data.pin
                sale.serial = data.username
                sale.amount = prices.get(type)
                sale.belongs = "Outreach"
                sale.free = free
                sale.year = year
                sale.deleted = 0

                sale.save()

                Aspnetusers.objects.filter(pin=data.pin).update(sold=1)

            return render(request, "sale_of_forms/bulk_outreach_receipt.html", context={'data': voucher})


@login_required
def download_excel_report(request):
    year = request.POST.get('year')
    programme = request.POST.get('programme')

    programmeData = Programmemodel.objects.all().order_by("type")

    if str(programmeData) != "":

        # print(programmeData.query)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = "attachment; filename=Admissions_statistics.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active

        # set_border(ws, "A4:D181")

        style_1 = Font(color='0000FF00', bold=True, size=15)

        ws.title = "Admissions_statistics"

        ws['A1'] = "TAKORADI TECHNICAL UNIVERSITY"
        ws['A2'] = "ADMISSIONS"
        ws['A3'] = "ADMISSIONS STATISTICS"

        row_num = 3

        columns = [
            (u"PROGRAMME", 60,),
            (u"2017/2018", 30),
            (u"2018/2019", 30),
            (u"2019/2020", 30),
            (u"2020/2021", 30),
            (u"2021/2022", 30),

        ]

        monthExam = ""
        yearExam = ""

        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            # c.style.font.bold = True
            # set column width
            ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

        for obj in programmeData:
            row_num += 1

            dataSet = [
                obj.name,
                totalAdmittedByYear("2017/2018", obj.id),
                totalAdmittedByYear("2018/2019", obj.id),
                totalAdmittedByYear("2019/2020", obj.id),
                totalAdmittedByYear("2020/2021", obj.id),
                totalAdmittedByYear("2021/2022", obj.id),

            ]
            # remove_duplicates(dataSet)
            for col_num in range(len(dataSet)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)

                c.value = dataSet[col_num]
                # c.style.alignment.wrap_text = True
                # c.style.font=style_1

        # formating

        sheet = wb.get_sheet_by_name("Admissions_statistics")
        sheet.sheet_properties.tabColor = "0072BA"

        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")

        border = Border(top=double, left=thin, right=thin, bottom=double)
        fill = PatternFill("solid", fgColor="DDDDDD")
        fill = GradientFill(stop=("000000", "FFFFFF"))
        font = Font(b=True, color="FF0000")
        al = Alignment(horizontal="center", vertical="center")

        wb.save(response)

        export_xlsx.short_description = u"Export results for verification XLSX"
        return response

    else:

        return render(request, 'dashboard/index.html')


@login_required
def export_xlsx(request):
    year = request.POST.get('year')
    programme = request.POST.get('programme')

    programmeData = Programmemodel.objects.filter(id=programme).get()

    programmeName = programmeData.name

    if year != "" and programme != "":

        data = Resultuploadmodel.objects.filter(applicant__yearofadmission=year).filter(
            applicant__admitted=True).filter(
            applicant__programmeadmittedid=programme).order_by('applicant__lastname').values('indexno',
                                                                                             'applicant__lastname',
                                                                                             'applicant__firstname',
                                                                                             'year',
                                                                                             'month',
                                                                                             'applicant__applicationnumber').annotate(
            total=Count('month'))
        print(data.query)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = "attachment; filename=" + programmeName + "_ " + year + "_admitted_applicants" + ".xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active

        # ws.cell('A1:A3').border = Border(top=Side(border_style='thin', color='FF000000'),
        #                               right=Side(border_style='thin', color='FF000000'),
        #                               bottom=Side(border_style='thin', color='FF000000'),
        #                               left=Side(border_style='thin', color='FF000000'))

        style_1 = Font(color='0000FF00', bold=True, size=15)

        ws.title = "Admitted Applicants"

        ws['A1'] = "TAKORADI TECHNICAL UNIVERSITY"
        ws['A2'] = "ADMISSIONS"
        ws['A3'] = "RESULT VERIFICATION FOR " + programmeName + "- " + str(year) + " ADMISSIONS "

        row_num = 3

        columns = [
            (u"NAME", 60,),
            (u"ADMISSION NUMBER", 30),
            (u"WEAC INDEX NUMBER", 30),
            (u"YEAR OF EXAMS", 30),
            (u"MONTH", 15),
        ]

        monthExam = ""
        yearExam = ""

        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            # c.style.font.bold = True
            # set column width
            ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

        for obj in data:
            row_num += 1

            if year in ["2018/2019", "2017/2018"]:

                if "MAY/JUNE" in obj['month'] or "may/june" in obj['month'] or "june" in obj['month'] or "may" in obj[
                    'month'] or "MayJune" in obj['month'] or "March" in obj['month'] or "/June" in obj[
                    'month'] or "/May" in \
                        obj['month'] or "April" in obj['month'] or "June/" in obj['month'] or "May/" in obj[
                    'month'] or "MAY/" in obj['month'] or "JUNE/" in obj['month'] or "JUNE" in obj['month'] or "MAY" in \
                        obj[
                            'month']:
                    monthExam = 6
                else:
                    monthExam = 12

                yearExam = str(obj['month']).rpartition(' ')[2]

                if not yearExam:
                    yearExam = str(obj['month']).rpartition(',')[2]

            else:
                monthExam = obj['month']

                yearExam = obj['year']
                # if not yearExam:
            #
            #     yearExam = obj.month.rpartition(',')[2]

            dataSet = [
                str(obj['applicant__lastname'] + ' ' + obj['applicant__firstname']).upper(),
                obj['applicant__applicationnumber'],
                obj['indexno'],
                yearExam,
                monthExam
            ]
            # remove_duplicates(dataSet)
            for col_num in range(len(dataSet)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)

                c.value = dataSet[col_num]
                # c.style.alignment.wrap_text = True
                # c.style.font=style_1

        # formating

        sheet = wb.get_sheet_by_name("Admitted Applicants")
        sheet.sheet_properties.tabColor = "0072BA"

        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")

        border = Border(top=double, left=thin, right=thin, bottom=double)
        fill = PatternFill("solid", fgColor="DDDDDD")
        fill = GradientFill(stop=("000000", "FFFFFF"))
        font = Font(b=True, color="FF0000")
        al = Alignment(horizontal="center", vertical="center")

        wb.save(response)

        export_xlsx.short_description = u"Export results for verification XLSX"
        return response

    else:

        return render(request, 'dashboard/index.html')


def getWeacMonth(applicant):
    data = Resultuploadmodel.objects.filter(applicant=applicant).values_list("month").annotate(
        total=Count("month"))

    return data


def graphs(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=graphs.xlsx'

    book = Workbook()
    sheet = book.active

    rows = [
        ("USA", 46),
        ("China", 38),
        ("UK", 29),
        ("Russia", 22),
        ("South Korea", 13),
        ("Germany", 11)
    ]

    for row in rows:
        sheet.append(row)

    ws2 = book.create_sheet("Graphs")

    data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=6)
    categs = Reference(sheet, min_col=1, min_row=1, max_row=6)

    chart = BarChart()
    chart.add_data(data=data)
    chart.set_categories(categs)

    chart.legend = None
    chart.y_axis.majorGridlines = None
    chart.varyColors = True
    chart.title = "Olympic Gold medals in London"

    ws2.add_chart(chart, "A8")

    # book.save("bar_chart.xlsx")

    book.save(response)
    return response


@login_required
def show_export_form(request):
    try:

        programme = Programmemodel.objects.filter(runing=True).filter(~Q(type="BTECH")).filter(
            ~Q(type="POSTGRADUATE")).order_by('name').all()

        year = ["2021/2022", "2020/2021", "2019/2020", "2018/2019", "2017/2018", "2016/2017", "2015/2016"]

        context = {'programme': programme, 'years': year}

        return render(request, 'reports/verification.html', context)
    except:

        messages.warning(request,
                         "Error displaying Applicant ")


def remove_duplicates(values):
    output = []
    seen = set()
    for value in values:
        # If value has not been encountered yet,
        # ... add it to both list and set.
        if value not in seen:
            output.append(value)
            seen.add(value)
    return output


def refire_sms(request, object_id):
    current_user = request.user

    applicant = Applicantmodel.objects.filter(applicationnumber=object_id).first()
    print(applicant.firstname)
    formDetails = getVoucher(applicant.applicationnumber)

    regular = applicant.firstname + "You have been admitted to TTU. Use this link; " \
                                    "admissions.ttuportal.com to print your admission letter." + str(
        formDetails)

    conditional = applicant.firstname + "You have been offered a conditional admission to TTU. Use " \
                                        "this link admissions.ttuportal.com to print your admission " \
                                        "letter." + str(
        formDetails)
    technical = applicant.firstname + "You have been admitted to TTU. Use this link; " \
                                      "admissions.ttuportal.com to print your admission letter." + str(
        formDetails)
    provisional = applicant.firstname + "You have been offered a provisional admission to TTU. Use " \
                                        "this link admissions.ttuportal.com to print your admission " \
                                        "letter." + str(
        formDetails)
    mature = applicant.firstname + "Your application for admission to TTU as mature student. Use this " \
                                   "link admissions.ttuportal.com to print your admission letter." + \
             str(
                 formDetails)
    email = applicant.email
    print(current_user)
    firesms(applicant.phone, regular, current_user.id, applicant.applicationnumber, request)
    if applicant.admissiontype == 'regular':
        firesms(applicant.phone, regular, current_user.id, applicant.applicationnumber, request)
        print(firesms(applicant.phone, regular, current_user.id, applicant.applicationnumber, request))
        # email = EmailMessage('Takoradi Technical University- Admissions Office', regular, to=[email])
        # email.send()
        subject = "Takoradi Technical University - Admissions"
        to = [applicant.email]
        from_email = 'admissions@ttu.edu.gh'

        ctx = {
            'applicant': applicant,
            'gad': ''

        }

        message = get_template('letters/2021/regular.html').render(ctx)

        msg = EmailMessage(subject, message, to=to, from_email=from_email)

        msg.content_subtype = 'html'

        msg.send()

        return redirect('applicants_index')

    elif applicant.admissiontype == 'mature':
        firesms(applicant.phone, mature, current_user.id, applicant.application_number, request)
        subject = "Takoradi Technical University - Admissions"
        to = [applicant.email]
        from_email = 'admissions@ttu.edu.gh'

        ctx = {
            'applicant': applicant,
            'gad': ''

        }

        message = get_template('letters/2021/regular.html').render(ctx)

        msg = EmailMessage(subject, message, to=to, from_email=from_email)

        msg.content_subtype = 'html'

        msg.send()

        return redirect('applicants_index')

    elif applicant.admissiontype == 'conditional':
        firesms(applicant.phone, conditional, current_user.id, applicant.application_number, request)
        # email = EmailMessage('Takoradi Technical University - Admissions Office', conditional, to=[email])
        # email.send()
    elif applicant.admissiontype == 'technical':
        firesms(applicant.phone, technical, current_user.id, applicant.application_number, request)
        subject = "Takoradi Technical University - Admissions"
        to = [applicant.email]
        from_email = 'admissions@ttu.edu.gh'

        ctx = {
            'applicant': applicant,
            'gad': ''

        }

        message = get_template('letters/2021/technical_email_.html').render(ctx)

        msg = EmailMessage(subject, message, to=to, from_email=from_email)

        msg.content_subtype = 'html'

        msg.send()

        return redirect('applicants_index')

    else:
        subject = "Takoradi Technical University - Admissions"
        to = [applicant.email]
        from_email = 'admissions@ttu.edu.gh'

        ctx = {
            'applicant': applicant,
            'gad': ''

        }

        message = get_template('letters/2021/regular.html').render(ctx)

        msg = EmailMessage(subject, message, to=to, from_email=from_email)

        msg.content_subtype = 'html'

        msg.send()

        return HttpResponse('email_two')


@login_required
def rooms(request):
    return render(request, 'accommodation/room.html')
